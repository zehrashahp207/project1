from flask import Blueprint, request, jsonify, render_template
import re, time
from pathlib import Path
from ipaddress import ip_address
from openpyxl import Workbook, load_workbook   # EXCEL kitabxanası

contact_bp = Blueprint("contact_bp", __name__, template_folder="templates")

PRIMARY = "#2A314D"
EXCEL_PATH = Path("messages.xlsx")   # EXCEL fayl yolu
last_submit_by_ip = {}

# REGEX-lər
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
FIRST_NAME_RE = re.compile(r"^[A-ZƏİÖÜÇŞĞ][a-zəiöüçşğ]+$")
LAST_NAME_RE = re.compile(r"^[A-ZƏİÖÜÇŞĞ][a-zəiöüçşğ]+$")

# Excel-ə mesaj yazmaq
def save_message_excel(first_name, last_name, email, message, ip):
    # Əgər fayl yoxdursa — yaradırıq və başlıqları yazırıq
    if not EXCEL_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"

        ws.append(["first_name", "last_name", "email", "message", "ip", "created_at"])
        wb.save(EXCEL_PATH)

    # Mövcud Excel faylını açıb yeni sətir əlavə edirik
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    ws.append([
        first_name,
        last_name,
        email,
        message,
        str(ip),
        time.strftime("%Y-%m-%dT%H:%M:%SZ", time.gmtime())
    ])

    wb.save(EXCEL_PATH)


# Validasiya
def validate_payload(data: dict):
    errors = {}
    first_name = (data.get("first_name") or "").strip()
    last_name = (data.get("last_name") or "").strip()
    email = (data.get("email") or "").strip()
    message = (data.get("message") or "").strip()
    hp = (data.get("hp") or "").strip()

    if not FIRST_NAME_RE.match(first_name):
        errors["first_name"] = "Ad düzgün deyil."
    if not LAST_NAME_RE.match(last_name):
        errors["last_name"] = "Soyad düzgün deyil."
    if not EMAIL_RE.match(email):
        errors["email"] = "Email düzgün deyil."
    if not (10 <= len(message) <= 2000):
        errors["message"] = "Mesaj 10–2000 simvol olmalıdır."
    if hp:
        errors["hp"] = "Bot şübhəsi (honeypot doludur)."

    return errors


# Əlaqə səhifəsi
@contact_bp.get("/əlaqə")
def əlaqə():
    return render_template("contact.html", primary=PRIMARY)


# API — mesaj göndərmək
@contact_bp.post("/api/əlaqə")
def api_əlaqə():
    data = request.get_json(force=True, silent=True) or {}
    errors = validate_payload(data)

    if errors:
        return jsonify({"error": "; ".join(f"{k}: {v}" for k, v in errors.items())}), 400

    # IP limit (anti-spam)
    try:
        client_ip = request.headers.get("X-Forwarded-For", request.remote_addr).split(",")[0].strip()
        client_ip = ip_address(client_ip)
    except Exception:
        client_ip = "0.0.0.0"

    now = time.time()
    last = last_submit_by_ip.get(client_ip)

    if last and (now - last) < 15:
        return jsonify({"error": "Çox tez-tez göndərirsiniz. 15 saniyə sonra yenidən cəhd edin."}), 429

    last_submit_by_ip[client_ip] = now

    # Mesaj EXCEL-ə yazılır ✔
    save_message_excel(
        first_name=data["first_name"].strip(),
        last_name=data["last_name"].strip(),
        email=data["email"].strip(),
        message=data["message"].strip(),
        ip=client_ip
    )

    return jsonify({"ok": True})


# Admin paneli — Excel-dən mesajları oxumaq
@contact_bp.get("/admin/messages")
def admin_messages():
    messages = []

    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]

        for row in ws.iter_rows(min_row=2, values_only=True):
            messages.append(dict(zip(headers, row)))

    return render_template("admin_messages.html", messages=messages, primary=PRIMARY)
