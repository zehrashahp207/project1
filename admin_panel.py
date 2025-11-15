from flask import Blueprint, render_template, redirect, url_for, flash, session, request
import json, os
from flask import send_file
import csv
from pathlib import Path
from openpyxl import load_workbook

MESSAGES_FILE = Path("messages.xlsx") 

admin_bp = Blueprint("admin_bp", __name__, template_folder="templates")

RESERV_FILE = "reservations.json"


# ======== Köməkçi funksiyalar ========
def load_reservations():
    if not os.path.exists(RESERV_FILE):
        return []
    with open(RESERV_FILE, "r", encoding="utf-8") as f:
        try:
            return json.load(f)
        except json.JSONDecodeError:
            return []


def save_reservations(reservations):
    with open(RESERV_FILE, "w", encoding="utf-8") as f:
        json.dump(reservations, f, ensure_ascii=False, indent=4)


# ======== ADMIN PANEL ========
# Admin paneldən mesajlara baxmaq
@admin_bp.route("/admin-panel/messages")
def admin_messages():
    messages = []

    if MESSAGES_FILE.exists():
        wb = load_workbook(MESSAGES_FILE)
        ws = wb.active

        # Başlıqları oxumaq üçün 2-ci sətirdən başlayırıq
        for row in ws.iter_rows(min_row=2, values_only=True):
            messages.append({
                "first_name": row[0],
                "last_name": row[1],
                "email": row[2],
                "message": row[3],
                "ip": row[4],
                "created_at": row[5]
            })

    return render_template("admin_messages.html", messages=messages)

@admin_bp.route("/admin-panel")
def admin_panel():
    if "user" not in session or session.get("role") != "admin":
        flash("Bu səhifəyə yalnız admin girə bilər!")
        return redirect(url_for("login"))

    reservations = load_reservations()

    pending = [r for r in reservations if r.get("status") in (None, "pending")]
    approved = [r for r in reservations if r.get("status") == "active"]
    rejected = [r for r in reservations if r.get("status") == "rejected"]

    return render_template(
        "admin_panel.html",
        pending=pending,
        approved=approved,
        rejected=rejected
    )


# ======== REZERVASİYANIN TƏSDİQİ ========
@admin_bp.route("/admin-panel/approve/<int:index>", methods=["POST"])
def approve_reservation(index):
    if "user" not in session or session.get("role") != "admin":
        flash("Yalnız admin təsdiqləyə bilər!")
        return redirect(url_for("login"))

    reservations = load_reservations()
    if 0 <= index < len(reservations):
        reservations[index]["status"] = "active"
        save_reservations(reservations)
        flash("Rezervasiya təsdiqləndi ✅")
    else:
        flash("Rezerv tapılmadı!")

    return redirect(url_for("admin_bp.admin_panel"))


# ======== REZERVASİYANIN RƏDD EDİLMƏSİ ========
@admin_bp.route("/admin-panel/reject/<int:index>", methods=["POST"])
def reject_reservation(index):
    if "user" not in session or session.get("role") != "admin":
        flash("Yalnız admin rədd edə bilər!")
        return redirect(url_for("login"))

    reservations = load_reservations()
    if 0 <= index < len(reservations):
        reservations[index]["status"] = "rejected"
        save_reservations(reservations)
        flash("Rezervasiya rədd edildi ❌")
    else:
        flash("Rezerv tapılmadı!")

    return redirect(url_for("admin_bp.admin_panel"))
